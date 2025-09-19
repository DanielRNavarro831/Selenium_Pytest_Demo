import openpyxl


class DocReader:
    doc_filename = "Selenium_Pytest_Data.xlsx"
    doc_path = f".//test_data//{doc_filename}"

    @staticmethod
    def get_cell_value_string(page: str, row_name: str):
        book = openpyxl.load_workbook(DocReader.doc_path)  # Opens the excel file
        sheet = book[page]  # gets the tab with the page name in the opened excel file
        cell_value = ""
        for a in range(1, sheet.max_row + 1):
            if sheet.cell(row=a, column=1).value == row_name:
                cell_value = sheet.cell(row=a, column=2).value  # sets "message" to the text in the string
                break
        return cell_value

    @staticmethod
    def get_new_email(current_email: str):
        split_email = current_email.split("@")  # ["MyEmail123", "Gmail.com"]
        digit = ""  # "Becomes "123"
        email_no_domain = split_email[0]  # "MyEmail123"
        email_no_digit = ""  # Becomes "MyEmail"
        for a in range(len(email_no_domain)):
            if email_no_domain[a].isdigit():
                digit += email_no_domain[a]
            else:
                email_no_digit += email_no_domain[a]
        new_digit = int(digit) + 1
        new_email = f"{email_no_digit}{new_digit}@{split_email[1]}"  # "MyEmail124@Gmail.com"
        return new_email

    @staticmethod
    def update_registration_email_doc():
        book = openpyxl.load_workbook(DocReader.doc_path)  # Opens the excel file
        sheet = book["Registration"]  # gets the tab with the page name in the opened excel file
        for a in range(1, sheet.max_row + 1):
            if sheet.cell(row=a, column=1).value == "Registration Email":
                current_email = sheet.cell(row=a, column=2).value
                new_email = DocReader.get_new_email(current_email)
                sheet.cell(row=a, column=2).value = new_email
                book.save(DocReader.doc_path)
                break
